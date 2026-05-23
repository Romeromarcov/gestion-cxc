import io
import logging
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from database import get_con
from routers.auth import get_current_user, require_roles
from routers.ventas import get_odoo
from routers.precios import precio_con_lista
from models.schemas import rows_to_list
from services.tasas_cambio import tasa_bcv_hoy, tasa_custom_hoy, tasa_bcv_en_fecha

logger = logging.getLogger(__name__)

router = APIRouter(prefix='/reportes', tags=['reportes'])


def _get_config(con) -> dict:
    try:
        rows = con.execute("SELECT clave, valor FROM config_app").fetchall()
        return {r['clave']: r['valor'] for r in rows}
    except Exception:
        return {}


def _nc_deductions(con, order_name: str) -> dict:
    """Calcula el valor de NCs aprobadas, desglosado por lista VES y lista USD.
    La asignación depende de la moneda dominante de los pagos de la orden.
    """
    notas = rows_to_list(con.execute(
        "SELECT id FROM notas_credito WHERE odoo_order_name=? AND estado='aprobada'",
        (order_name,)
    ).fetchall())
    if not notas:
        return {'nc_lista_ves': 0.0, 'nc_lista_usd': 0.0}

    pagos = rows_to_list(con.execute("""
        SELECT moneda, SUM(monto) as total FROM pagos
        WHERE odoo_order_name=?
          AND estado IN ('recibido','enviado_odoo','confirmado_odoo','enviado_sheets')
        GROUP BY moneda
    """, (order_name,)).fetchall())
    monto_ves = sum(p['total'] for p in pagos if p['moneda'] == 'VES')
    monto_usd = sum(p['total'] for p in pagos if p['moneda'] in ('USD', 'USDT'))

    replica = con.execute(
        "SELECT id FROM ordenes_replica WHERE odoo_order_name=? AND estado='activa'",
        (order_name,)
    ).fetchone()
    precio_usd_map = {}
    if replica:
        for r in con.execute(
            "SELECT odoo_line_id, precio_lista_usd FROM ordenes_replica_lineas WHERE replica_id=?",
            (replica['id'],)
        ).fetchall():
            precio_usd_map[r['odoo_line_id']] = r['precio_lista_usd']

    nc_ves = 0.0
    nc_usd = 0.0
    for nota in notas:
        for l in rows_to_list(con.execute("""
            SELECT odoo_line_id, cantidad, precio_original, descuento_aprobado
            FROM notas_credito_lineas WHERE nota_id=?
        """, (nota['id'],)).fetchall()):
            if l.get('descuento_aprobado') is None:
                continue
            qty = float(l.get('cantidad') or 1)
            pct = float(l['descuento_aprobado']) / 100.0
            val_ves = qty * float(l.get('precio_original') or 0) * pct
            val_usd = (qty * float(precio_usd_map.get(l['odoo_line_id']) or 0) * pct
                       if l.get('odoo_line_id') and l['odoo_line_id'] in precio_usd_map
                       else val_ves)
            nc_ves += val_ves
            nc_usd += val_usd

    if monto_ves > 0 and monto_usd == 0:
        return {'nc_lista_ves': nc_ves, 'nc_lista_usd': 0.0}
    elif monto_usd > 0 and monto_ves == 0:
        return {'nc_lista_ves': 0.0, 'nc_lista_usd': nc_usd}
    elif monto_ves > 0 and monto_usd > 0:
        t = monto_ves + monto_usd
        return {'nc_lista_ves': nc_ves * (monto_ves / t),
                'nc_lista_usd': nc_usd * (monto_usd / t)}
    return {'nc_lista_ves': nc_ves, 'nc_lista_usd': 0.0}


def _calcular_estado_cobro(con, order_name: str,
                            sub_lista_ves, tax_lista_ves, tot_lista_ves,
                            sub_lista_usd, tax_lista_usd, tot_lista_usd,
                            moneda_orden: str, nc: dict,
                            tolerancia: float, facturas_odoo: list) -> dict:
    """Estado de cobro unificado pre/post factura.

    Estados posibles:
      pendiente        — sin pagos aún
      pago_parcial     — pagos, pero no alcanzan el subtotal
      subtotal_cobrado — pagos ≈ subtotal (listo para facturar + retención)
      pagado_completo  — pagos ≈ total (sin retención pendiente)
      sobrepagado      — pagos > total + tolerancia
      en_proceso_odoo  — factura publicada, pago parcial en Odoo
      cerrado          — factura publicada y amount_residual ≈ 0

    POST-FACTURA (facturas_odoo no vacío):
      Odoo es la fuente de verdad vía amount_residual.
      Incluye automáticamente retenciones, pagos directos y cualquier
      crédito conciliado contra la factura.
      Se suman también pagos del sistema aún no enviados a Odoo.
    """
    pagos = rows_to_list(con.execute("""
        SELECT monto, moneda, fecha_pago FROM pagos
        WHERE odoo_order_name=?
          AND estado IN ('recibido','enviado_odoo','confirmado_odoo','enviado_sheets')
        ORDER BY fecha_pago
    """, (order_name,)).fetchall())

    par = 'EUR_VES' if (moneda_orden or '').upper() == 'EUR' else 'USD_VES'

    def _usd(p_list):
        s = 0.0
        for p in p_list:
            m = float(p['monto'] or 0)
            mon = p['moneda']
            if mon in ('USD', 'USDT', 'EUR'):
                s += m
            elif mon == 'VES':
                t = tasa_bcv_en_fecha(par, p.get('fecha_pago') or '')
                if t:
                    s += m / t
        return s

    def _ves(p_list):
        s = 0.0
        for p in p_list:
            m = float(p['monto'] or 0)
            mon = p['moneda']
            if mon == 'VES':
                s += m
            elif mon in ('USD', 'USDT', 'EUR'):
                t = tasa_bcv_en_fecha(par, p.get('fecha_pago') or '')
                if t:
                    s += m * t
        return s

    suma_usd = _usd(pagos)
    suma_ves = _ves(pagos)
    monedas  = {p['moneda'] for p in pagos}
    todos_ves = bool(monedas) and monedas <= {'VES'}
    todos_usd = bool(monedas) and monedas <= {'USD', 'USDT', 'EUR'}

    nc_ves = float(nc.get('nc_lista_ves') or 0)
    nc_usd = float(nc.get('nc_lista_usd') or 0)

    obj_tot_ves = max(0.0, (tot_lista_ves or 0) - nc_ves)
    obj_tot_usd = max(0.0, (tot_lista_usd or 0) - nc_usd)
    obj_sub_ves = max(0.0, (sub_lista_ves or 0))
    obj_sub_usd = max(0.0, (sub_lista_usd or 0))

    tol_ves = tolerancia * (tasa_bcv_hoy(par) or 1)

    # ── POST-FACTURA: Odoo como fuente de verdad ─────────────────────────────
    if facturas_odoo:
        total_facturado = sum(float(f.get('amount_total') or 0) for f in facturas_odoo)
        total_residual  = sum(float(f.get('amount_residual') or 0) for f in facturas_odoo)
        cobrado_odoo    = total_facturado - total_residual

        # Pagos del sistema que aún no llegaron a Odoo
        pagos_no_odoo = rows_to_list(con.execute("""
            SELECT monto, moneda, fecha_pago FROM pagos
            WHERE odoo_order_name=? AND estado='recibido' AND odoo_payment_id IS NULL
        """, (order_name,)).fetchall())
        pendiente_envio_usd = _usd(pagos_no_odoo)

        total_cobrado_combinado = cobrado_odoo + pendiente_envio_usd

        pagado_completo = total_residual <= tolerancia
        sobrepagado     = total_residual < -tolerancia
        sobrepago_usd   = abs(total_residual) if sobrepagado else 0.0

        if pagado_completo:
            estado = 'cerrado'
        elif sobrepagado:
            estado = 'sobrepagado'
        elif total_cobrado_combinado > 0:
            estado = 'en_proceso_odoo'
        else:
            estado = 'pendiente'

        return {
            'estado_cobro': estado,
            'pagado_completo': pagado_completo,
            'subtotal_cobrado': False,
            'sobrepagado': sobrepagado,
            'sobrepago_usd': sobrepago_usd,
            'has_invoice': True,
            'cobrado_odoo': cobrado_odoo,
            'residual_odoo': total_residual,
            'pendiente_envio_usd': pendiente_envio_usd,
            'suma_usd_sistema': suma_usd,
            'suma_ves_sistema': suma_ves,
            # compat
            'pagado_lista_ves': False,
            'pagado_lista_usd': pagado_completo,
            'saldo_lista_ves': None,
            'saldo_lista_usd': total_residual,
        }

    # ── PRE-FACTURA: pagos del sistema vs listas ─────────────────────────────
    pagado_tot_ves = (todos_ves and obj_tot_ves > 0 and
                      abs(suma_ves - obj_tot_ves) <= tol_ves)
    pagado_tot_usd = (todos_usd and obj_tot_usd > 0 and
                      abs(suma_usd - obj_tot_usd) <= tolerancia)
    pagado_completo = pagado_tot_ves or pagado_tot_usd

    # Subtotal cobrado: ≈ subtotal pero < total (para retención)
    sub_cobrado_ves = (todos_ves and obj_sub_ves > 0 and not pagado_completo and
                       abs(suma_ves - obj_sub_ves) <= tol_ves)
    sub_cobrado_usd = (todos_usd and obj_sub_usd > 0 and not pagado_completo and
                       abs(suma_usd - obj_sub_usd) <= tolerancia)
    subtotal_cobrado = sub_cobrado_ves or sub_cobrado_usd

    # Sobrepago
    sobre_ves = (todos_ves and obj_tot_ves > 0 and
                 suma_ves > obj_tot_ves + tol_ves)
    sobre_usd = (todos_usd and obj_tot_usd > 0 and
                 suma_usd > obj_tot_usd + tolerancia)
    sobrepagado = sobre_ves or sobre_usd

    sobrepago_usd = 0.0
    if sobre_usd and obj_tot_usd > 0:
        sobrepago_usd = suma_usd - obj_tot_usd
    elif sobre_ves and obj_tot_ves > 0:
        t = tasa_bcv_hoy(par)
        sobrepago_usd = (suma_ves - obj_tot_ves) / t if t else 0.0

    # Estado final
    if sobrepagado:
        estado = 'sobrepagado'
    elif pagado_completo:
        estado = 'pagado_completo'
    elif subtotal_cobrado:
        estado = 'subtotal_cobrado'
    elif pagos:
        estado = 'pago_parcial'
    else:
        estado = 'pendiente'

    return {
        'estado_cobro': estado,
        'pagado_completo': pagado_completo,
        'subtotal_cobrado': subtotal_cobrado,
        'sobrepagado': sobrepagado,
        'sobrepago_usd': sobrepago_usd,
        'has_invoice': False,
        'cobrado_odoo': None,
        'residual_odoo': None,
        'pendiente_envio_usd': 0.0,
        'suma_usd_sistema': suma_usd,
        'suma_ves_sistema': suma_ves,
        'pagado_lista_ves': pagado_tot_ves,
        'pagado_lista_usd': pagado_tot_usd,
        'saldo_lista_ves': obj_tot_ves - suma_ves if todos_ves else None,
        'saldo_lista_usd': obj_tot_usd - suma_usd if todos_usd else None,
    }


@router.get('/cxc')
def reporte_cxc(lista_id: int = None,
                skip: int = 0, limit: int = Query(default=200, le=1000),
                user=Depends(require_roles('gerente', 'admin'))):
    """Reporte CxC con estado de cobro completo: pre/post factura, retención, sobrepago."""
    odoo = get_odoo()
    tasa_bcv    = tasa_bcv_hoy() or 0
    tasa_custom = tasa_custom_hoy() or 0

    try:
        ventas_odoo = odoo.get_ventas()
    except Exception as e:
        logger.warning('reportes: get_ventas falló — %s', e)
        ventas_odoo = []

    con = get_con()
    cfg = _get_config(con)
    tolerancia    = float(cfg.get('tolerancia_pago_usd', '0.01') or 0.01)
    pl_usd_nombre = cfg.get('pricelist_usd_nombre', 'Lista USD')
    pl_ves_nombre = cfg.get('pricelist_ves_nombre', 'Precio USD Pago VES')

    # Batch: pagos del sistema agrupados por orden
    pagos_rows = rows_to_list(con.execute("""
        SELECT odoo_order_name, SUM(equivalente_usd) as total_usd
        FROM pagos
        WHERE estado IN ('recibido','enviado_sheets','enviado_odoo','confirmado_odoo')
          AND odoo_order_name IS NOT NULL
        GROUP BY odoo_order_name
    """).fetchall())
    pagos_por_orden = {p['odoo_order_name']: p['total_usd'] for p in pagos_rows}

    # Batch: réplicas activas
    replicas_map = {r['odoo_order_name']: r for r in rows_to_list(con.execute("""
        SELECT odoo_order_name, moneda_orden,
               subtotal_lista_ves, tax_lista_ves, total_lista_ves,
               subtotal_lista_usd, tax_lista_usd, total_lista_usd
        FROM ordenes_replica WHERE estado='activa'
    """).fetchall())}

    # Batch: facturas publicadas en Odoo (una sola llamada)
    try:
        order_names    = [v['name'] for v in ventas_odoo]
        facturas_map   = odoo.get_facturas_por_orden(order_names)
    except Exception as e:
        logger.warning('reportes: get_facturas_por_orden falló — %s', e)
        facturas_map   = {}

    resultado = []
    for v in ventas_odoo:
        order_name = v['name']

        # Montos Odoo
        subtotal_odoo = float(v.get('amount_untaxed') or 0)
        tax_odoo      = float(v.get('amount_tax') or 0)
        total_odoo    = float(v.get('amount_total') or 0)
        pagado_usd    = float(pagos_por_orden.get(order_name) or 0)

        # Moneda / lista de la orden en Odoo
        moneda_orden = (v['currency_id'][1]
                        if isinstance(v.get('currency_id'), list) else 'USD')
        pricelist_odoo = (v['pricelist_id'][1]
                          if isinstance(v.get('pricelist_id'), list) else '')

        replica = replicas_map.get(order_name)

        # Columnas de lista
        if replica:
            sub_ves = replica.get('subtotal_lista_ves')
            tax_ves = replica.get('tax_lista_ves')
            tot_ves = replica.get('total_lista_ves')
            sub_usd = replica.get('subtotal_lista_usd')
            tax_usd = replica.get('tax_lista_usd')
            tot_usd = replica.get('total_lista_usd')
            moneda_orden = replica.get('moneda_orden') or moneda_orden
        elif pricelist_odoo == pl_usd_nombre:
            sub_ves, tax_ves, tot_ves = None, None, None
            sub_usd = subtotal_odoo
            tax_usd = tax_odoo
            tot_usd = total_odoo
        else:
            sub_ves = tax_ves = tot_ves = None
            sub_usd = tax_usd = tot_usd = None

        nc = _nc_deductions(con, order_name)

        facturas_orden = facturas_map.get(order_name, [])
        cobro = _calcular_estado_cobro(
            con, order_name,
            sub_ves, tax_ves, tot_ves,
            sub_usd, tax_usd, tot_usd,
            moneda_orden, nc, tolerancia, facturas_orden
        )

        # Saldo: post-factura usa residual de Odoo (exacto); pre-factura usa diferencia simple
        if cobro['has_invoice'] and cobro['residual_odoo'] is not None:
            saldo_usd = cobro['residual_odoo']
        else:
            saldo_usd = total_odoo - pagado_usd

        # Estado CxC legacy (para compatibilidad con el campo anterior)
        ec = cobro['estado_cobro']
        if ec in ('cerrado', 'pagado_completo'):
            estado_cxc = 'pagado'
        elif ec == 'sobrepagado':
            estado_cxc = 'sobrepagado'
        elif ec in ('subtotal_cobrado', 'en_proceso_odoo', 'pago_parcial'):
            estado_cxc = 'parcialmente_pagado'
        else:
            estado_cxc = 'pendiente'

        resultado.append({
            'codigo': order_name,
            'partner_id': v['partner_id'][0] if v.get('partner_id') else None,
            'cliente': v['partner_id'][1] if v.get('partner_id') else '',
            'fecha': v.get('date_order', ''),
            'estado': v.get('state'),
            'invoice_status': v.get('invoice_status'),
            'moneda_orden': moneda_orden,
            'pricelist_odoo': pricelist_odoo,
            # Montos Odoo
            'subtotal_odoo': subtotal_odoo,
            'tax_odoo': tax_odoo,
            'total_odoo': total_odoo,
            'pagado_usd': pagado_usd,
            'saldo_usd': saldo_usd,
            'saldo_ves_bcv': saldo_usd * tasa_bcv,
            # Lista VES (solo con réplica)
            'subtotal_lista_ves': sub_ves,
            'tax_lista_ves':      tax_ves,
            'total_lista_ves':    tot_ves,
            'nc_lista_ves':       nc['nc_lista_ves'],
            'a_pagar_lista_ves':  ((tot_ves - nc['nc_lista_ves']) if tot_ves is not None else None),
            # Lista USD (réplica o directo desde Odoo)
            'subtotal_lista_usd': sub_usd,
            'tax_lista_usd':      tax_usd,
            'total_lista_usd':    tot_usd,
            'nc_lista_usd':       nc['nc_lista_usd'],
            'a_pagar_lista_usd':  ((tot_usd - nc['nc_lista_usd']) if tot_usd is not None else None),
            # Estado de cobro
            'estado_cxc': estado_cxc,
            'estado_cobro': ec,
            'pagado_lista_ves': cobro.get('pagado_lista_ves'),
            'pagado_lista_usd': cobro.get('pagado_lista_usd'),
            'subtotal_cobrado': cobro['subtotal_cobrado'],
            'sobrepagado': cobro['sobrepagado'],
            'sobrepago_usd': cobro.get('sobrepago_usd', 0),
            # Facturas en Odoo
            'has_invoice': cobro['has_invoice'],
            'cobrado_odoo': cobro.get('cobrado_odoo'),
            'residual_odoo': cobro.get('residual_odoo'),
            'pendiente_envio_usd': cobro.get('pendiente_envio_usd', 0),
            'tiene_replica': replica is not None,
            'lista_usd_directo': pricelist_odoo == pl_usd_nombre,
            'facturas_odoo': [
                {'nombre': f.get('name'), 'total': f.get('amount_total'),
                 'residual': f.get('amount_residual'),
                 'payment_state': f.get('payment_state')}
                for f in facturas_orden
            ],
        })

    con.close()
    total_saldo = sum(r['saldo_usd'] for r in resultado)
    return {
        'tasa_bcv': tasa_bcv,
        'tasa_custom': tasa_custom,
        'total_saldo_usd': total_saldo,
        'total_saldo_ves_bcv': total_saldo * tasa_bcv,
        'total': len(resultado),
        'ventas': resultado[skip:skip + limit],
    }


@router.get('/alertas')
def get_alertas(user=Depends(require_roles('gerente', 'admin'))):
    """Conteos rápidos desde DB local (sin llamadas a Odoo) para el dashboard."""
    con = get_con()
    cfg = _get_config(con)
    tol = float(cfg.get('tolerancia_pago_usd', '0.01') or 0.01)

    # Sobrepago: sum(equivalente_usd) > total_lista_usd + tolerancia
    sobrepago = con.execute("""
        SELECT COUNT(*) FROM (
            SELECT p.odoo_order_name
            FROM pagos p
            JOIN ordenes_replica r ON r.odoo_order_name = p.odoo_order_name
                                  AND r.estado = 'activa'
            WHERE p.estado IN ('recibido','enviado_odoo','confirmado_odoo','enviado_sheets')
              AND p.equivalente_usd IS NOT NULL
              AND r.total_lista_usd IS NOT NULL
            GROUP BY p.odoo_order_name
            HAVING SUM(p.equivalente_usd) > r.total_lista_usd + ?
        )
    """, (tol,)).fetchone()[0]

    # Subtotal cobrado: sum(equiv_usd) entre [sub-tol, sub+tol] y < total-tol
    sub_cobrado = con.execute("""
        SELECT COUNT(*) FROM (
            SELECT p.odoo_order_name
            FROM pagos p
            JOIN ordenes_replica r ON r.odoo_order_name = p.odoo_order_name
                                  AND r.estado = 'activa'
            WHERE p.estado IN ('recibido','enviado_odoo','confirmado_odoo','enviado_sheets')
              AND p.equivalente_usd IS NOT NULL
              AND r.subtotal_lista_usd IS NOT NULL AND r.total_lista_usd IS NOT NULL
            GROUP BY p.odoo_order_name
            HAVING ABS(SUM(p.equivalente_usd) - r.subtotal_lista_usd) <= ?
               AND SUM(p.equivalente_usd) < r.total_lista_usd - ?
        )
    """, (tol, tol)).fetchone()[0]

    creditos_disponibles = con.execute(
        "SELECT COUNT(*) FROM creditos_cliente WHERE estado='disponible'"
    ).fetchone()[0]

    con.close()
    return {
        'sobrepago_count': sobrepago,
        'subtotal_cobrado_count': sub_cobrado,
        'creditos_disponibles': creditos_disponibles,
    }


@router.get('/ordenes-pagadas-pendientes')
def ordenes_pagadas_pendientes(user=Depends(require_roles('gerente', 'admin'))):
    """Órdenes cuyo cobro está completo (o subtotal cobrado) pero factura no publicada."""
    odoo = get_odoo()
    con = get_con()
    cfg = _get_config(con)
    tolerancia = float(cfg.get('tolerancia_pago_usd', '0.01') or 0.01)

    replicas = rows_to_list(con.execute("""
        SELECT odoo_order_name, moneda_orden,
               subtotal_lista_ves, tax_lista_ves, total_lista_ves,
               subtotal_lista_usd, tax_lista_usd, total_lista_usd
        FROM ordenes_replica WHERE estado='activa'
    """).fetchall())

    try:
        facturas_map = odoo.get_facturas_por_orden(
            [r['odoo_order_name'] for r in replicas])
    except Exception:
        facturas_map = {}

    resultado = []
    for replica in replicas:
        order_name = replica['odoo_order_name']
        nc = _nc_deductions(con, order_name)
        cobro = _calcular_estado_cobro(
            con, order_name,
            replica['subtotal_lista_ves'], replica['tax_lista_ves'], replica['total_lista_ves'],
            replica['subtotal_lista_usd'], replica['tax_lista_usd'], replica['total_lista_usd'],
            replica['moneda_orden'], nc, tolerancia,
            facturas_map.get(order_name, [])
        )
        if cobro['estado_cobro'] not in ('subtotal_cobrado', 'pagado_completo', 'cerrado'):
            continue
        if cobro['estado_cobro'] == 'cerrado':
            continue  # factura ya publicada y pagada, no alertar
        resultado.append({
            'orden': order_name,
            'estado_cobro': cobro['estado_cobro'],
            'total_lista_usd': replica['total_lista_usd'],
            'cobrado_usd': cobro['suma_usd_sistema'],
            'has_invoice': cobro['has_invoice'],
        })

    con.close()
    return resultado


@router.get('/ventas')
def reporte_ventas(vendedor_id: int = None, cliente: str = None,
                   fecha_desde: str = None, fecha_hasta: str = None,
                   marca: str = None, lista_id: int = None,
                   skip: int = 0, limit: int = Query(default=200, le=1000),
                   user=Depends(require_roles('gerente', 'admin'))):
    odoo = get_odoo()
    try:
        ventas_odoo = odoo.get_ventas()
    except Exception as e:
        logger.warning('reportes: get_ventas falló — %s', e)
        ventas_odoo = []

    con = get_con()
    extras = {r['producto_ref']: dict(r) for r in rows_to_list(
        con.execute("SELECT * FROM productos_extra").fetchall())}
    lista_info = None
    if lista_id:
        lista_info = con.execute(
            "SELECT * FROM listas_precios WHERE id=?", (lista_id,)
        ).fetchone()

    resultado = []
    for v in ventas_odoo:
        if fecha_desde and v.get('date_order', '') < fecha_desde:
            continue
        if fecha_hasta and v.get('date_order', '') > fecha_hasta:
            continue
        if cliente and cliente.lower() not in (
                v['partner_id'][1].lower() if v.get('partner_id') else ''):
            continue
        for l in odoo.get_lineas_venta(v['id']):
            ref = l.get('default_code') or ''
            extra = extras.get(ref, {})
            if marca and extra.get('marca', '').lower() != marca.lower():
                continue
            precio_lista = None
            if lista_info:
                precio_lista = precio_con_lista(
                    l, lista_id,
                    lista_info['umbral_descuento_excluir'] if lista_info else None)
            resultado.append({
                'orden': v['name'],
                'cliente': v['partner_id'][1] if v.get('partner_id') else '',
                'fecha': v.get('date_order', ''),
                'producto_ref': ref,
                'producto': l['product_id'][1] if l.get('product_id') else '',
                'cantidad': l.get('product_uom_qty'),
                'precio_odoo': l.get('price_unit'),
                'descuento_pct': l.get('discount', 0),
                'subtotal': l.get('price_subtotal'),
                'precio_lista': precio_lista,
                'marca': extra.get('marca'),
                'categoria_local': extra.get('categoria_local'),
            })
    con.close()
    return resultado[skip:skip + limit]


@router.get('/ventas/exportar-excel')
def exportar_excel(user=Depends(require_roles('gerente', 'admin'))):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail='openpyxl no instalado')

    odoo = get_odoo()
    try:
        ventas = odoo.get_ventas()
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas CxC'
    headers = ['Orden', 'Cliente', 'Fecha', 'Subtotal USD', 'IVA', 'Total USD',
               'Estado', 'Facturación']
    hf = PatternFill('solid', fgColor='2E75B6')
    hfont = Font(bold=True, color='FFFFFF')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont
        c.alignment = Alignment(horizontal='center')
    for ri, v in enumerate(ventas, 2):
        ws.cell(ri, 1, v.get('name'))
        ws.cell(ri, 2, v['partner_id'][1] if v.get('partner_id') else '')
        ws.cell(ri, 3, v.get('date_order', ''))
        ws.cell(ri, 4, v.get('amount_untaxed', 0))
        ws.cell(ri, 5, v.get('amount_tax', 0))
        ws.cell(ri, 6, v.get('amount_total', 0))
        ws.cell(ri, 7, v.get('state'))
        ws.cell(ri, 8, v.get('invoice_status'))
    for col in ws.columns:
        w = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, w + 2)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=ventas_cxc.xlsx'})


@router.get('/resumen-dashboard')
def resumen_dashboard(user=Depends(require_roles('gerente', 'admin'))):
    from datetime import date
    con = get_con()
    total_propuestos = con.execute(
        "SELECT COUNT(*) FROM notas_credito WHERE estado='enviada'"
    ).fetchone()[0]
    total_pagos_pendientes = con.execute(
        "SELECT COUNT(*) FROM pagos WHERE estado='propuesto'"
    ).fetchone()[0]
    total_pagos_recibidos = con.execute(
        "SELECT COUNT(*) FROM pagos WHERE estado='recibido'"
    ).fetchone()[0]
    replicas_con_error = con.execute(
        "SELECT COUNT(*) FROM ordenes_replica WHERE estado='error'"
    ).fetchone()[0]
    creditos_disponibles = con.execute(
        "SELECT COUNT(*) FROM creditos_cliente WHERE estado='disponible'"
    ).fetchone()[0]

    # Gastos recurrentes y servicios públicos pendientes de pago este mes
    hoy = date.today()
    mes_actual = hoy.strftime('%Y-%m')
    gastos_recurrentes_pendientes = con.execute("""
        SELECT COUNT(*) FROM gastos
        WHERE tipo IN ('recurrente','servicio_publico')
          AND estado IN ('borrador','pendiente')
          AND strftime('%Y-%m', fecha) <= ?
    """, (mes_actual,)).fetchone()[0]

    # CxP: total monto pendiente a proveedores (compras sin pagar en maestro)
    cxp_row = con.execute("""
        SELECT ROUND(COALESCE(SUM(monto_usd_bcv), 0), 2)
        FROM maestro_operaciones
        WHERE tipo='egreso'
          AND origen IN ('odoo_auto_proveedor','zelle_terceros')
          AND estado='registrado'
    """).fetchone()
    cxp_total_usd = cxp_row[0] if cxp_row else 0

    pendientes_aprobacion = (
        con.execute("SELECT COUNT(*) FROM gastos WHERE aprobacion_estado='pendiente'").fetchone()[0] +
        con.execute("SELECT COUNT(*) FROM cambios_divisa WHERE aprobacion_estado='pendiente'").fetchone()[0] +
        con.execute("SELECT COUNT(*) FROM compras_internas WHERE aprobacion_estado='pendiente'").fetchone()[0]
    )

    con.close()
    return {
        'aprobaciones_pendientes': total_propuestos,
        'pagos_pendientes_recepcion': total_pagos_pendientes,
        'pagos_listos_para_odoo': total_pagos_recibidos,
        'replicas_con_error': replicas_con_error,
        'creditos_disponibles': creditos_disponibles,
        'gastos_recurrentes_pendientes': gastos_recurrentes_pendientes,
        'cxp_total_usd': float(cxp_total_usd),
        'pendientes_aprobacion': pendientes_aprobacion,
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL CONSOLIDADO
# ─────────────────────────────────────────────────────────────────────────────

def _wb_header(ws, headers: list, color: str = '2E75B6'):
    """Escribe encabezados con fondo azul (o el color indicado) en la hoja ws."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail='openpyxl no instalado')
    fill  = PatternFill('solid', fgColor=color)
    font  = Font(bold=True, color='FFFFFF')
    align = Alignment(horizontal='center', vertical='center')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill; c.font = font; c.alignment = align
    ws.row_dimensions[1].height = 18


def _wb_autowidth(ws):
    """Ajusta ancho de columnas al contenido."""
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)


@router.get('/exportar-excel-consolidado')
def exportar_excel_consolidado(
    desde: str = None,
    hasta: str = None,
    user=Depends(require_roles('gerente', 'admin'))
):
    """
    Descarga un .xlsx con 5 hojas:
      1. Maestro de Operaciones
      2. Gastos
      3. Nómina
      4. Pagos Fiscales
      5. Resumen
    Acepta filtros ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail='openpyxl no instalado')

    con = get_con()
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)           # eliminar hoja default vacía

    from datetime import date
    hoy    = date.today().isoformat()
    f_desde = desde or '2000-01-01'
    f_hasta = hasta or hoy

    # ── 1. Maestro de Operaciones ──────────────────────────────────────────
    ws1 = wb.create_sheet('Maestro')
    _wb_header(ws1, [
        'ID', 'Fecha', 'Tipo', 'Descripción', 'Categoría', 'Subcategoría',
        'Método', 'Moneda', 'Monto', 'Tasa BCV', 'Monto USD (BCV)',
        'Origen', 'Referencia Odoo', 'Estado', 'Nro Documento'
    ], color='1F4E79')
    rows = con.execute("""
        SELECT id, fecha, tipo, descripcion, categoria, subcategoria,
               metodo, moneda, monto, tasa_bcv, monto_usd_bcv,
               origen, odoo_ref, estado, nro_documento
        FROM maestro_operaciones
        WHERE fecha BETWEEN ? AND ?
        ORDER BY fecha DESC
    """, (f_desde, f_hasta)).fetchall()
    for ri, r in enumerate(rows, 2):
        for ci, val in enumerate(r, 1):
            ws1.cell(ri, ci, val)
    _wb_autowidth(ws1)

    # ── 2. Gastos ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Gastos')
    _wb_header(ws2, [
        'ID', 'Fecha', 'Tipo', 'Categoría', 'Subcategoría', 'Descripción',
        'Monto', 'Moneda', 'Estado', 'Es Recurrente', 'Período Pago',
        'Proveedor Odoo', 'Move Odoo', 'Maestro ID'
    ], color='375623')
    rows = con.execute("""
        SELECT id, fecha, tipo, categoria, subcategoria, descripcion,
               monto, moneda, estado, es_recurrente, periodo_pago,
               odoo_partner_id, odoo_move_id, maestro_op_id
        FROM gastos
        WHERE fecha BETWEEN ? AND ?
        ORDER BY fecha DESC
    """, (f_desde, f_hasta)).fetchall()
    for ri, r in enumerate(rows, 2):
        for ci, val in enumerate(r, 1):
            ws2.cell(ri, ci, val)
    _wb_autowidth(ws2)

    # ── 3. Nómina ──────────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Nómina')
    _wb_header(ws3, [
        'ID', 'Período', 'Tipo', 'Descripción', 'Monto', 'Moneda',
        'Tasa BCV', 'Monto USD', 'Estado', 'Origen',
        'Move Odoo', 'Maestro ID', 'Creado En'
    ], color='7B3F00')
    rows = con.execute("""
        SELECT id, periodo, tipo, descripcion, monto, moneda,
               tasa_bcv, monto_usd, estado, origen,
               odoo_move_id, maestro_op_id, creado_en
        FROM nomina_registros
        WHERE (periodo BETWEEN ? AND ? OR creado_en BETWEEN ? AND ?)
        ORDER BY periodo DESC
    """, (f_desde[:7], f_hasta[:7], f_desde, f_hasta)).fetchall()
    for ri, r in enumerate(rows, 2):
        for ci, val in enumerate(r, 1):
            ws3.cell(ri, ci, val)
    _wb_autowidth(ws3)

    # ── 4. Pagos Fiscales ──────────────────────────────────────────────────
    ws4 = wb.create_sheet('Pagos Fiscales')
    _wb_header(ws4, [
        'ID', 'Fecha', 'Tipo', 'Descripción', 'Monto', 'Moneda',
        'Tasa BCV', 'Monto USD', 'Banco/Ref', 'Estado',
        'Move Odoo', 'Maestro ID'
    ], color='4A235A')
    rows = con.execute("""
        SELECT id, fecha, tipo, descripcion, monto, moneda,
               tasa_bcv, monto_usd, banco_ref, estado,
               odoo_move_id, maestro_op_id
        FROM pagos_fiscales
        WHERE fecha BETWEEN ? AND ?
        ORDER BY fecha DESC
    """, (f_desde, f_hasta)).fetchall()
    for ri, r in enumerate(rows, 2):
        for ci, val in enumerate(r, 1):
            ws4.cell(ri, ci, val)
    _wb_autowidth(ws4)

    # ── 5. Resumen ─────────────────────────────────────────────────────────
    ws5 = wb.create_sheet('Resumen')
    _wb_header(ws5, ['Concepto', 'Valor'], color='2E75B6')
    # totales rápidos
    tot_ingreso = con.execute(
        "SELECT ROUND(COALESCE(SUM(monto_usd_bcv),0),2) FROM maestro_operaciones "
        "WHERE tipo='ingreso' AND fecha BETWEEN ? AND ?", (f_desde, f_hasta)
    ).fetchone()[0] or 0
    tot_egreso = con.execute(
        "SELECT ROUND(COALESCE(SUM(monto_usd_bcv),0),2) FROM maestro_operaciones "
        "WHERE tipo='egreso' AND fecha BETWEEN ? AND ?", (f_desde, f_hasta)
    ).fetchone()[0] or 0
    tot_gastos = con.execute(
        "SELECT ROUND(COALESCE(SUM(monto),0),2) FROM gastos "
        "WHERE moneda='USD' AND estado='pagado' AND fecha BETWEEN ? AND ?", (f_desde, f_hasta)
    ).fetchone()[0] or 0
    tot_nomina = con.execute(
        "SELECT ROUND(COALESCE(SUM(monto_usd),0),2) FROM nomina_registros "
        "WHERE estado IN ('enviado_odoo','descontado') "
        "AND creado_en BETWEEN ? AND ?", (f_desde, f_hasta)
    ).fetchone()[0] or 0
    tot_fiscal = con.execute(
        "SELECT ROUND(COALESCE(SUM(monto_usd),0),2) FROM pagos_fiscales "
        "WHERE estado='validado' AND fecha BETWEEN ? AND ?", (f_desde, f_hasta)
    ).fetchone()[0] or 0
    cnt_req = con.execute(
        "SELECT COUNT(*) FROM requisiciones "
        "WHERE creado_en BETWEEN ? AND ?", (f_desde, f_hasta)
    ).fetchone()[0] or 0

    resumen = [
        ('Período analizado', f'{f_desde}  →  {f_hasta}'),
        ('', ''),
        ('Total Ingresos USD (Maestro)',  float(tot_ingreso)),
        ('Total Egresos USD (Maestro)',   float(tot_egreso)),
        ('Balance neto USD',              round(float(tot_ingreso) - float(tot_egreso), 2)),
        ('', ''),
        ('Gastos pagados USD',            float(tot_gastos)),
        ('Nómina ejecutada USD',          float(tot_nomina)),
        ('Pagos fiscales validados USD',  float(tot_fiscal)),
        ('Requisiciones creadas',         int(cnt_req)),
        ('', ''),
        ('Generado el', hoy),
    ]
    for ri, (concepto, valor) in enumerate(resumen, 2):
        ws5.cell(ri, 1, concepto)
        ws5.cell(ri, 2, valor)
        if concepto in ('Balance neto USD',):
            from openpyxl.styles import Font
            ws5.cell(ri, 2).font = Font(bold=True,
                color='375623' if float(valor) >= 0 else 'C00000')
    _wb_autowidth(ws5)
    con.close()

    nombre_archivo = f'consolidado_{f_desde}_{f_hasta}.xlsx'
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={nombre_archivo}'}
    )
